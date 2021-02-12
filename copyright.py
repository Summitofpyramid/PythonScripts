import requests
import tarfile
import openpyxl
import subprocess

#The script it assumes there will be a list of packages with
# url in a sheet, then it download the package and extract,
# detect license file, and save the license type to the sheet
def getCopyRight():
    urlIndex = 13
    copyrightIndex= 12
    original = '/Users/qhe/PycharmProjects/pythonProject/Book3.xlsx'

    book = openpyxl.load_workbook(original)
    sheet = book.worksheets[0]

    for index in range(1, sheet.max_row):
        if sheet.cell(row=index, column=urlIndex).value is None:
            print("row ", index, 'has no url')
            continue
        print("row: ", index, sheet.cell(row=index, column=urlIndex).value)
        res = requests.get(sheet.cell(row=index, column=urlIndex).value, allow_redirects=True)
        index_of_last = sheet.cell(row=index, column=urlIndex).value.rfind('/')
        filename = sheet.cell(row=index, column=urlIndex).value[index_of_last + 1:]
        foldername = filename[0:filename.rfind('.')]
        open('./packages/' + filename, 'wb').write(res.content)
        Tarfile = tarfile.open('./packages/' + filename, 'r')
        Tarfile.extractall('./extractedPackages/' + foldername)
        out = subprocess.Popen(['find', './extractedPackages/' + foldername, '-iname', '*LICENSE*'],
                               stdout=subprocess.PIPE,
                               stderr=subprocess.STDOUT)
        stdout, stderr = out.communicate()
        licensepath = stdout.decode('utf-8')
        if licensepath is None or len(licensepath) == 0:
            print('No license file found!')
            continue
        else:
            print('License path: ', licensepath)
            # grep LICENSE file with pattern 'Copyright (c)'
            licensegrep = subprocess.Popen('grep Copyright '+licensepath, shell=True,
                                           stdout=subprocess.PIPE,
                                           stderr=subprocess.STDOUT)
            stdout, stderr = licensegrep.communicate()
            print(stdout.decode('utf-8'))
            sheet.cell(index, copyrightIndex).value = stdout.decode('utf-8')
    book.save('Book3.xlsx')
