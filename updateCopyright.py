import requests
import tarfile
import openpyxl
import subprocess

def findIndex(pacakgeName:str, sheet, col:int):
    for i in range(1,sheet.max_row):
        if pacakgeName == sheet.cell(row=i, column=col).value:
            return i
    return -1

def updateSpreadsheet():

    # updateSheetPath = "/Users/qhe/PycharmProjects/pythonProject/ANA793520200830 - WMLA 2.2 Auth Module Review Level 2 Bulk 238 Open Source Tracking Sheet.xlsx"
    # searchItemsPath = "/Users/qhe/PycharmProjects/pythonProject/WMLA 2.2 Auth Module Scan Report (ANA793920200901 Packages requiring no pedigree review).xlsx"
    updateSheetPath = "/Users/qhe/PycharmProjects/pythonProject/2update.xlsx"
    searchItemsPath = "/Users/qhe/PycharmProjects/pythonProject/2search.xlsx"
    probesPath = "/Users/qhe/PycharmProjects/pythonProject/Book2.xlsx"

    updateBook = openpyxl.load_workbook(updateSheetPath)
    updateSheet = updateBook.worksheets[0]

    searchBook = openpyxl.load_workbook(searchItemsPath)
    searchSheet = searchBook.worksheets[0]

    probeBook = openpyxl.load_workbook(probesPath)
    probeSheet = probeBook.worksheets[0]

    for i in range(1,probeSheet.max_row+1):
        indexSearch = findIndex(probeSheet.cell(row=i,column=1).value, searchSheet,4)
        indexUpdate = findIndex(probeSheet.cell(row=i,column=1).value, updateSheet,4)
        if indexSearch == -1 or indexUpdate == -1:
            print(probeSheet.cell(row=i,column=1).value, "not found!")
        if searchSheet.cell(indexSearch,14).value is None:
            print(probeSheet.cell(row=i,column=1).value)
        print("------------------------------")
        updateSheet.cell(indexUpdate,12).value = searchSheet.cell(indexSearch,14).value

    updateBook.save(updateSheetPath)